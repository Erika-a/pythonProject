import openpyxl
from openpyxl.styles import PatternFill
import os
import time

# file1 = openpyxl.load_workbook("C:\\Users\\PC\\Desktop\\地铁二号线百度坐标.xlsx")

# print(file1.sheetnames)
#
# sheet = file1['站点']
#
# print(type(sheet))
#
# print(sheet.title)
#
# anothersheet = file1.active
#
# print(anothersheet)
#
# print(sheet['A1'])
#
# print(sheet['A1'].value)
#
# c = sheet['A2']
# print(c.value)
#
# print("行{0}, 列{1} 是 {2}".format(c.row,c.column,c.value))
#
# print('cell {0} is {1}'.format(c.coordinate,c.value))
#
# print(sheet['A3'].value)
#
# print(sheet.cell(row=1,column=2))
#
# print(sheet.cell(row=1,column=2).value)
#
# for i in range(1,8,1):
#     print(i,sheet.cell(row=i,column=1).value + "\t" + sheet.cell(row=i,column=3).value)
#
# print(sheet.max_column,sheet.max_row)

# 1.遍历找出相同的站名
# 2.
filepath1 = "C:\\Users\\PC\\Desktop\\兰州城市学院B2.xlsx"
filepath2 = "C:\\Users\\PC\\Desktop\\GDJTNBSP1.xlsx"

file1 = openpyxl.load_workbook(filepath1)
file2 = openpyxl.load_workbook(filepath2)

# print(file1.sheetnames)

sheet = file1['Summary']
sheet2 = file2['GDJTNBSP']

# print(sheet['C3'].value)
# print(sheet2['B8'].value)
#
# print(sheet.max_column,sheet.max_row)
# print(sheet2.max_column,sheet2.max_row)
list1 = ''
yuanx = 0
yuany = 0
xianx = 0
xiany = 0
zhongx = 0
zhongy = 0

fillB1 = PatternFill("solid", fgColor="CCE5FF")
fillB2 = PatternFill("solid", fgColor="CCFFCC")
fillB3 =PatternFill("solid", fgColor="FFCCCC")

filename1 = filepath1.split('\\')[-1][:-7]


filepath3 = filepath1.replace('B1','B2')
filepath4 = filepath1.replace('B1','B3')


if os.path.exists(filepath3):
    print('ok')



for i in range(2, sheet2.max_row):
    b = sheet2.cell(row=i, column=1).value
    list1 = list1 + b + ','

alist = list1.split(',')

for i in range(2, sheet.max_row):
    c = sheet.cell(row=i, column=5).value
    b = sheet.cell(row=i, column=7).value
    if c != None:
        yuany = yuany + c
        yuanx = yuanx + sheet.cell(row=i, column=4).value
        xianx = xianx + float(sheet.cell(row=i, column=1).value)
        xiany = xiany + float(sheet.cell(row=i, column=2).value)
        zhongx = yuanx - xianx
        zhongy = yuany - xiany
    elif b != None:
        zhongx = sheet.cell(row=i, column=7).value
        zhongy = sheet.cell(row=i, column=8).value


# zhongx = format(zhongx, '.2f')
# zhongy = format(zhongy, '.2f')

print(yuanx, yuany)
print(xianx, xiany)
print(zhongx, zhongy)



# wb = openpyxl.Workbook("C:\\Users\\PC\\Desktop\\东方红广场站cctv.xlsx")
# file1.create_sheet(index=1, title='zonghe')
# 
# sheet3 = file1['zonghe']

for i in range(1,sheet.max_row+1):
    val1 = sheet.cell(row=i, column=3).value
    for r in range(1,sheet2.max_row+1):
        val2 = sheet2.cell(row=r, column=1).value
        set1 = sheet2.cell(row=r,column=2).value
        if val1 in val2 and filename1 in set1:
            valx1 = format((float(sheet.cell(row=i, column=1).value) + float(zhongx)), '.2f')
            valy1 = format((float(sheet.cell(row=i, column=2).value) + float(zhongy)), '.2f')
            valx2 = sheet2.cell(row=r, column=9).value
            valy2 = sheet2.cell(row=r, column=10).value
            print(val2,valx1,valy1)
            # sheet.cell(row=i, column=7).value = zhongx
            # sheet.cell(row=i, column=8).value = zhongy
            sheet2.cell(row=r, column=9).value = valx1
            sheet2.cell(row=r, column=10).value = valy1

            if 'B1' in str(filepath1):
                sheet2.cell(row=r, column=9).fill = fillB1
                sheet2.cell(row=r, column=10).fill = fillB1
            elif 'B2' in str(filepath1):
                sheet2.cell(row=r, column=9).fill = fillB2
                sheet2.cell(row=r, column=10).fill = fillB2
            else:
                sheet2.cell(row=r, column=9).fill = fillB3
                sheet2.cell(row=r, column=10).fill = fillB3


if os.path.exists(filepath3):
    filex = openpyxl.load_workbook(filepath3)
    sheetx = filex['Summary']
    sheetx.cell(row=2, column=7).value = zhongx
    sheetx.cell(row=2, column=8).value = zhongy
    filex.save(filepath3)
elif os.path.exists(filepath4):
    filey = openpyxl.load_workbook(filepath4)
    sheety = filey['Summary']
    sheety.cell(row=2, column=7).value = zhongx
    sheety.cell(row=2, column=8).value = zhongy
    filey.save(filepath4)

# #
# for i in range(2, sheet.max_row):
#     val1 = sheet.cell(row=i, column=3).value
#
#     for r in alist:
#         if val1 in r:
#             val2 = alist.index(r)
#             set2 = sheet2.cell(row=val2 + 2, column=2).value
#
#             if filename1 in set2:
#                 valx1 = format((float(sheet.cell(row=i, column=1).value) + float(zhongx)), '.2f')
#                 valy1 = format((float(sheet.cell(row=i, column=2).value) + float(zhongy)), '.2f')
#                 valx2 = sheet2.cell(row=val2 + 2, column=9).value
#                 valy2 = sheet2.cell(row=val2 + 2, column=10).value
#                 print(r + '\t\t' + str(valx1) + '\t\t' + str(valy1))
#                 sheet2.cell(row=val2 + 2, column=9).value = valx1
#                 sheet2.cell(row=val2 + 2, column=10).value = valy1
#                 if 'B1' in str(filepath1):
#                     sheet2.cell(row=val2 + 2, column=9).fill = fillB1
#                     sheet2.cell(row=val2 + 2, column=10).fill = fillB1
#                 elif 'B2' in str(filepath1):
#                     sheet2.cell(row=val2 + 2, column=9).fill = fillB2
#                     sheet2.cell(row=val2 + 2, column=10).fill = fillB2
#                 else:
#                     sheet2.cell(row=val2 + 2, column=9).fill = fillB3


file1.save(filepath1)

file2.save(filepath2)



# for i in alist:
#     if 'QJ-05' in i:
#         print(i + str(alist.index(i)))
#
# print(sheet2.cell(row=2+2,column=1).value)
