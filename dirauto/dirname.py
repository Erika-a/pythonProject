import os
import shutil
import xlsxwriter
import time
from xlsxwriter import workbook
import pandas as pd
import openpyxl



# 获取所有文件路径以及文件名 格式 类型 写入 excel


def getfiles():
    path = r'E:\Zhangjiakou_20210815\Zhangjiakou_20210815_files\3DT\2050图层'
    filelist1 = []
    filelist2 = []
    for dirpath, dirnames, filenames in os.walk(path):
        for filename in filenames:
            filelist1.append(dirpath.replace('E:\\Zhangjiakou_20210815\\Zhangjiakou_20210815_files\\', '').replace('\\',
                                                                                                                   '!') + '!' + filename)

    # print(filelist1)
    return filelist1


# print(f'{filename}\t\t,{dirnames}\t\t,{dirpath}\t\t')


def excel1():
    file = xlsxwriter.Workbook(
        filename='C:\\Users\\PC\\Desktop\\test2.xlsx',
        options={  # 全局设置
            'strings_to_numbers': True,  # str 类型数字转换为 int 数字
            'strings_to_urls': False,  # 自动识别超链接
            'constant_memory': False,  # 连续内存模式 (True 适用于大数据量输出)
            'default_format_properties': {
                'font_name': '微软雅黑',  # 字体. 默认值 "Arial"
                'font_size': 10,  # 字号. 默认值 11
                # 'bold': False,  # 字体加粗
                # 'border': 1,  # 单元格边框宽度. 默认值 0
                # 'align': 'left',  # 对齐方式
                # 'valign': 'vcenter',  # 垂直对齐方式
                # 'text_wrap': False,  # 单元格内是否自动换行
                # ...
            },
        }
    )
    my_format = file.add_format({
        'fg_color': '#c5d9f1',  # 单元格填充颜色. 常用的颜色可以用单词描述. 自定义的颜色可以用 '#FFFFFF' 这样描述
    })
    fmt = file.add_format({'align': 'center', 'valign': 'vcenter'})
    worksheet = file.add_worksheet('name')
    l = 2
    # c5d9f1
    for i in getfiles():
        h = 2
        i = i.split('!')
        for t in i:
            if '.' in t:
                worksheet.write(l, h, t)
                h = h + 1
            else:
                worksheet.write(l, h, t, my_format)
                h = h + 1
        l = l + 1
        print(i[-1] + "写入完毕-------")
    print('excel file ok')
    file.close()

def readexcel():
    openpyxl.reader('C:\\Users\\PC\\Desktop\\test2.xlsx')




def readex():
    df = pd.read_excel('C:\\Users\\PC\\Desktop\\test2.xlsx')
    data1 = df.head(7)  # 读取前7行的所有数据，dataFrame结构

    data2 = df.values  # list形式，读取表格所有数据

    print("获取到所有的值:\n{0}".format(data1))  # 格式化输出

    print("获取到所有的值:\n{0}".format(data2))  # 格式化输出


# # 单元格合并后居中
# fmt = book.add_format({'align': 'center', 'valign': 'vcenter'})
# # sheet.merge_range(x1, y1, x2, y2, value, cell_format=None)
# sheet.merge_range(0, 0, 1, 10, 'hello', cell_format=fmt)


#         worksheet.write(l,h,i)
#         l = l + 1
#         if l >= 20:
#             l = 0
#             h = h + 2
#         print(i+'写入完毕-------')
# print('excel file ok')
# file.close()


def excel2():
    file = xlsxwriter.Workbook('C:\\Users\\PC\\Desktop\\test.xlsx')

    worksheet = file.add_worksheet('name')
    l = 0
    h = 3

    for i in getfiles():
        if '.3dt' in i:
            worksheet.write(l, h, i)
            l = l + 1
            if l >= 20:
                l = 0
                h = h + 2
            print(i + '写入完毕-------')
    print('excel file ok')
    file.close()


if __name__ == '__main__':
    excel1()

# book = xlsxwriter.Workbook(
#     filename='./test.xlsx',
#     options={  # 全局设置
#         'strings_to_numbers': True,  # str 类型数字转换为 int 数字
#         'strings_to_urls': False,  # 自动识别超链接
#         'constant_memory': False,  # 连续内存模式 (True 适用于大数据量输出)
#         'default_format_properties': {
#             'font_name': '微软雅黑',  # 字体. 默认值 "Arial"
#             'font_size': 10,  # 字号. 默认值 11
#             # 'bold': False,  # 字体加粗
#             # 'border': 1,  # 单元格边框宽度. 默认值 0
#             # 'align': 'left',  # 对齐方式
#             # 'valign': 'vcenter',  # 垂直对齐方式
#             # 'text_wrap': False,  # 单元格内是否自动换行
#             # ...
#         },
#     }
# )


#
# import pandas as pd
#
#
# def set_foramt():
#   # 1.多个sheet页
#   sheet_list = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']
#   # 2.要写入的excel路径
#   writer = pd.ExcelWriter('test_new.xlsx')
#   workbook = writer.book
#   # 3.循环遍历sheet取数据
#   for sheet in sheet_list:
# 	  df1 = pd.read_excel('test.xlsx', sheet_name=sheet)
#
# 	  # 4.设置格式
# 	  fmt = workbook.add_format({"font_name": u"微软雅黑"})
# 	  percent_fmt = workbook.add_format({'num_format': '0.00%'})
# 	  amt_fmt = workbook.add_format({'num_format': '#,##0.00'})
# 	  border_format = workbook.add_format({'border': 1})
# 	  bg_format = workbook.add_format({'bold': True, 'font_name': u'微软雅黑', 'bg_color': 'yellow',
# 									   'align': 'center', 'valign': 'vcenter', 'font_color': 'black',
# 									   'font_size': 10})
# 	  date_fmt = workbook.add_format({'bold': False, 'font_name': u'微软雅黑', 'num_format': 'yyyy-mm-dd',
# 									  'align': 'center', 'valign': 'vcenter'})
# 	  # 5.写入excel
# 	  l_end = len(df1.index) + 1
# 	  df1.to_excel(writer, sheet_name=sheet, encoding='utf8', header=df1.columns.values.tolist(),
# 				   index=False, startcol=0, startrow=0)
# 	  worksheet1 = writer.sheets[sheet]
#
# 	  # 6.生效单元格格式
# 	  # 设置行高
# 	  worksheet1.set_row(0, 20, fmt)  # 从第0行开始，行高为20，格式为fmt
# 	  # 设置列宽
# 	  worksheet1.set_column('A:F', 20, fmt)  # 从A列到F列，行高为10，格式为fmt
# 	  # 加边框
# 	  worksheet1.conditional_format(f'A1:F{l_end}', {'type': 'no_blanks', 'format': border_format})
# 								  # A1单元格到F(索引值）, 'type': 'no_blanks'指非空的单元格加格式
# 	  # 设置背景色
# 	  worksheet1.conditional_format('A1:F1', {'type': 'no_blanks', 'format': bg_format})
# 	  worksheet1.conditional_format(f'B2:B{l_end}', {'type': 'no_blanks', 'format': bg_format})
# 	  # 合并日期单元格
# 	  worksheet1.merge_range(f'A2:A{l_end}', df1['日期'][0], date_fmt)
# 	  # 设置数值格式千分位，并保留两位小数
# 	  worksheet1.conditional_format(f'C2:F{l_end}', {'type': 'no_blanks', 'format': amt_fmt})
#   # 7.所有的sheet页设置好后再保存
